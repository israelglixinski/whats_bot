from openpyxl import load_workbook
import config



def gravar():
    arquivo_xlsx = load_workbook(f'{config.path_dir}/Dados.xlsx')
    planilha = arquivo_xlsx['Contatos']


    planilha['A1'] = 'Número'
    planilha['A2'] = '5541995789107'
    planilha['A3'] = '5541995789107'
    planilha['A4'] = '5541995789107'
    planilha['A5'] = '5541995789107'
    

    planilha['B1'] = 'Mensagem'
    planilha['B2'] = 'Teste1'
    planilha['B3'] = 'Teste2'
    planilha['B4'] = 'Teste3'
    planilha['B5'] = 'Teste4'
    

    planilha['C1'] = 'Status'
    planilha['C2'] = ''
    planilha['C3'] = ''
    planilha['C4'] = ''
    planilha['C5'] = ''
    

    planilha['D1'] = ''
    planilha['D2'] = ''
    planilha['D3'] = ''
    planilha['D4'] = ''
    planilha['D5'] = ''
    

    planilha['E1'] = ''
    planilha['E2'] = ''
    planilha['E3'] = ''
    planilha['E4'] = ''
    planilha['E5'] = ''
    

    arquivo_xlsx.save(f'{config.path_dir}/Dados.xlsx')


# https://web.whatsapp.com/send?phone=5541984988524&text=HI

def visualizar():

    for item in range(5):

        arquivo_xlsx = load_workbook(f'{config.path_dir}/Dados.xlsx')
        planilha = arquivo_xlsx['Contatos']
    
        linha = item + 1
        AAA = planilha[f'A{linha}'].value
        BBB = planilha[f'B{linha}'].value
        CCC = planilha[f'C{linha}'].value
        DDD = planilha[f'D{linha}'].value        
    
        print(f'{AAA}---{BBB}---{CCC}')



gravar()
visualizar()